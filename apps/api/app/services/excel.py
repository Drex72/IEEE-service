from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.models.schemas import CompanyContactImport, CompanyImportRow, TrackerSummary


EMAIL_PATTERN = re.compile(r"[\w.+-]+@[\w-]+\.[\w.-]+")
URL_PATTERN = re.compile(
    r"(?<!@)\b(?:https?://)?(?:www\.)?[A-Za-z0-9.-]+\.[A-Za-z]{2,}(?:/[^\s]*)?"
)
LINKEDIN_PATTERN = re.compile(r"https?://(?:[\w.]+\.)?linkedin\.com/[^\s)]+", re.IGNORECASE)
GENERIC_ROUTE_PATTERN = re.compile(
    r"\b(?:via|use|apply|reach|request|call|walk-?in|portal|contact form|website|linkedin)\b",
    re.IGNORECASE,
)
PERSON_STOPWORDS = {
    "via",
    "website",
    "linkedin",
    "contact",
    "form",
    "office",
    "team",
    "request",
    "reach",
    "call",
    "apply",
    "portal",
    "direct",
    "corporate",
    "foundation",
    "csr",
    "marketing",
    "communications",
    "customer",
    "care",
    "sales",
    "support",
    "online",
    "enquiries",
}


class ExcelSponsorTrackerParser:
    def parse(self, file_bytes: bytes) -> tuple[list[CompanyImportRow], TrackerSummary | None]:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        header_row_index, headers = self._locate_headers(sheet)
        rows: list[CompanyImportRow] = []
        primary_contact_seen: set[str] = set()

        for excel_row_index in range(header_row_index + 1, sheet.max_row + 1):
            row_values = [sheet.cell(row=excel_row_index, column=idx + 1).value for idx in range(len(headers))]
            normalized = {headers[idx]: self._stringify(value) for idx, value in enumerate(row_values)}
            company_name = normalized.get("company")
            if not company_name:
                continue
            if company_name.lower().startswith("status key"):
                continue

            contact_details = normalized.get("contact / email")
            contact_emails = self._extract_emails(contact_details)
            website = self._extract_website(
                contact_details,
                normalized.get("how to reach"),
                normalized.get("notes"),
            )
            company_key = company_name.strip().lower()
            contact = self._build_contact(
                company_name=company_name,
                raw_contact=contact_details,
                contact_email=contact_emails[0] if contact_emails else None,
                phone_or_address=normalized.get("phone / address"),
                reach_channel=normalized.get("how to reach"),
                notes=normalized.get("notes"),
                source_row=excel_row_index,
                is_primary=company_key not in primary_contact_seen,
            )
            primary_contact_seen.add(company_key)
            contact_summary = self._summarize_contact(contact, fallback=contact_details)

            rows.append(
                CompanyImportRow(
                    name=company_name,
                    website=website,
                    industry=normalized.get("category"),
                    tier=normalized.get("tier"),
                    contact_email=contact.email if contact else (contact_emails[0] if contact_emails else None),
                    contact_details=contact_summary,
                    phone_or_address=normalized.get("phone / address"),
                    reach_channel=normalized.get("how to reach"),
                    notes=normalized.get("notes"),
                    status=normalized.get("status"),
                    source_row=excel_row_index,
                    metadata={
                        "date_sent": normalized.get("date sent"),
                        "follow_up_date": normalized.get("follow-up date"),
                        "index": normalized.get("#"),
                        "raw_contact_cell": contact_details,
                    },
                    contacts=[contact] if contact else [],
                )
            )

        return rows, self._parse_summary_sheet(workbook)

    def _locate_headers(self, sheet: Any) -> tuple[int, list[str]]:
        for excel_row_index in range(1, min(sheet.max_row, 12) + 1):
            row_values = [self._stringify(cell.value) for cell in sheet[excel_row_index]]
            normalized_headers = [value.lower() if value else "" for value in row_values]
            if "company" in normalized_headers and "status" in normalized_headers:
                return excel_row_index, normalized_headers
        raise ValueError("Could not locate sponsor tracker headers.")

    def _parse_summary_sheet(self, workbook: Any) -> TrackerSummary | None:
        if len(workbook.sheetnames) < 2:
            return None

        sheet = workbook[workbook.sheetnames[1]]
        instructions: list[str] = []
        banner = self._stringify(sheet["A1"].value)
        tier_guide: str | None = None
        ieee_angle: str | None = None
        for row_index in range(2, sheet.max_row + 1):
            label = self._stringify(sheet.cell(row=row_index, column=1).value)
            detail = self._stringify(sheet.cell(row=row_index, column=2).value)
            if not label and not detail:
                continue
            if label and label.startswith("STEP") and detail:
                instructions.append(detail)
            if label == "TIER GUIDE":
                tier_guide = detail
            if label == "IEEE ANGLE":
                ieee_angle = detail

        return TrackerSummary(
            banner=banner,
            context_line=self._stringify(sheet["B2"].value),
            instructions=instructions,
            tier_guide=tier_guide,
            ieee_angle=ieee_angle,
        )

    def _extract_email(self, raw_value: str | None) -> str | None:
        emails = self._extract_emails(raw_value)
        return emails[0] if emails else None

    def _extract_emails(self, raw_value: str | None) -> list[str]:
        if not raw_value:
            return []
        seen: set[str] = set()
        emails: list[str] = []
        for match in EMAIL_PATTERN.findall(raw_value):
            normalized = match.lower()
            if normalized in seen:
                continue
            seen.add(normalized)
            emails.append(normalized)
        return emails

    def _extract_linkedin_url(self, *values: str | None) -> str | None:
        for value in values:
            if not value:
                continue
            match = LINKEDIN_PATTERN.search(value)
            if match:
                return match.group(0).strip(" .,/)")
        return None

    def _extract_urls(self, *values: str | None) -> list[str]:
        seen: set[str] = set()
        urls: list[str] = []
        for value in values:
            if not value:
                continue
            for match in URL_PATTERN.findall(value):
                candidate = match.strip(" /.,)")
                if "@" in candidate:
                    continue
                if not candidate.startswith(("http://", "https://")):
                    candidate = f"https://{candidate}"
                if candidate in seen:
                    continue
                seen.add(candidate)
                urls.append(candidate)
        return urls

    def _build_contact(
        self,
        *,
        company_name: str,
        raw_contact: str | None,
        contact_email: str | None,
        phone_or_address: str | None,
        reach_channel: str | None,
        notes: str | None,
        source_row: int,
        is_primary: bool,
    ) -> CompanyContactImport | None:
        if not any([raw_contact, contact_email, phone_or_address, reach_channel, notes]):
            return None

        cleaned_contact = raw_contact or contact_email
        emails = self._extract_emails(cleaned_contact)
        primary_email = contact_email or (emails[0] if emails else None)
        urls = self._extract_urls(cleaned_contact, reach_channel)
        linkedin_url = self._extract_linkedin_url(cleaned_contact, reach_channel, notes)
        non_linkedin_urls = [
            url for url in urls if "linkedin.com" not in url.lower()
        ]
        full_name = self._extract_contact_name(cleaned_contact, primary_email)
        role_title = self._extract_contact_role(cleaned_contact, full_name, primary_email)
        contact_kind, display_label = self._classify_contact_surface(
            raw_contact=cleaned_contact,
            primary_email=primary_email,
            full_name=full_name,
            linkedin_url=linkedin_url,
            non_linkedin_urls=non_linkedin_urls,
            reach_channel=reach_channel,
        )
        identity_bits = [
            company_name.strip().lower(),
            str(source_row),
            (primary_email or "").strip().lower(),
            (full_name or "").strip().lower(),
            (role_title or "").strip().lower(),
            contact_kind,
        ]
        external_key = "|".join(bit for bit in identity_bits if bit) or f"{company_name.lower()}|{source_row}"

        metadata: dict[str, Any] = {
            "contact_kind": contact_kind,
            "raw_contact_input": cleaned_contact,
        }
        if emails[1:]:
            metadata["additional_emails"] = emails[1:]
        if non_linkedin_urls:
            metadata["contact_urls"] = non_linkedin_urls
            metadata["contact_form_url"] = non_linkedin_urls[0]
        if reach_channel:
            metadata["route_hint"] = reach_channel
        if notes and contact_kind in {"contact_form", "route"}:
            metadata["notes_context"] = notes

        return CompanyContactImport(
            external_key=external_key,
            raw_contact=display_label,
            full_name=full_name,
            role_title=role_title,
            email=primary_email,
            linkedin_url=linkedin_url,
            phone_or_address=phone_or_address,
            reach_channel=reach_channel,
            notes=notes,
            source_row=source_row,
            metadata=metadata,
            is_primary=is_primary,
        )

    def _extract_contact_name(
        self,
        raw_contact: str | None,
        contact_email: str | None,
    ) -> str | None:
        if not raw_contact:
            return None
        cleaned = EMAIL_PATTERN.sub(" ", raw_contact)
        cleaned = URL_PATTERN.sub(" ", cleaned)
        cleaned = re.sub(r"[\[\]<>]", " ", cleaned)
        cleaned = re.sub(r"\b(?:email|mail|contact|reach)\b", " ", cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r"[|;/]+", ",", cleaned)
        candidate = re.split(r",|\(|\)|\n", cleaned, maxsplit=1)[0].strip(" -:")
        candidate = re.sub(r"\s{2,}", " ", candidate).strip()
        if not candidate:
            return None
        lowered = candidate.lower()
        if contact_email and lowered == contact_email.lower():
            return None
        candidate_words = [
            word.strip(".").lower()
            for word in re.split(r"\s+", candidate)
            if word.strip(".")
        ]
        if len(candidate_words) < 2 or len(candidate_words) > 5:
            return None
        if any(word in PERSON_STOPWORDS for word in candidate_words):
            return None
        if GENERIC_ROUTE_PATTERN.search(candidate):
            return None
        if not any(word[:1].isupper() for word in candidate.split() if word):
            return None
        return candidate

    def _extract_contact_role(
        self,
        raw_contact: str | None,
        full_name: str | None,
        contact_email: str | None,
    ) -> str | None:
        if not raw_contact:
            return None
        cleaned = EMAIL_PATTERN.sub(" ", raw_contact)
        cleaned = URL_PATTERN.sub(" ", cleaned)
        if full_name:
            cleaned = cleaned.replace(full_name, " ")
        if contact_email:
            cleaned = cleaned.replace(contact_email, " ")
        fragments = [fragment.strip(" -:|,/") for fragment in re.split(r",|\(|\)|\n", cleaned)]
        for fragment in fragments:
            if not fragment:
                continue
            lowered = fragment.lower()
            if lowered in {"email", "mail", "contact"} or GENERIC_ROUTE_PATTERN.search(fragment):
                continue
            if len(fragment.split()) <= 8:
                return fragment
        return None

    def _classify_contact_surface(
        self,
        *,
        raw_contact: str | None,
        primary_email: str | None,
        full_name: str | None,
        linkedin_url: str | None,
        non_linkedin_urls: list[str],
        reach_channel: str | None,
    ) -> tuple[str, str]:
        combined_text = " ".join(part for part in [raw_contact, reach_channel] if part).lower()
        mentions_linkedin = "linkedin" in combined_text or linkedin_url is not None
        mentions_website = "website" in combined_text or bool(non_linkedin_urls)
        mentions_contact_form = "contact form" in combined_text or any(
            "contact" in url.lower() for url in non_linkedin_urls
        )

        if full_name:
            return "person", full_name
        if primary_email:
            return "email", primary_email
        if mentions_linkedin and mentions_website:
            return "route", "LinkedIn or website outreach"
        if mentions_contact_form or non_linkedin_urls:
            return "contact_form", "Website contact form"
        if mentions_linkedin:
            return "linkedin", "LinkedIn outreach"
        if "call" in combined_text and "office" in combined_text:
            return "route", "Call office directly"
        return "route", "General outreach route"

    def _summarize_contact(
        self,
        contact: CompanyContactImport | None,
        *,
        fallback: str | None,
    ) -> str | None:
        if contact:
            if contact.full_name and contact.role_title:
                return f"{contact.full_name}, {contact.role_title}"
            if contact.full_name:
                return contact.full_name
            if contact.email:
                return contact.email
            if contact.raw_contact:
                return contact.raw_contact
        return fallback

    def _extract_website(self, *values: str | None) -> str | None:
        for value in values:
            if not value:
                continue
            for match in URL_PATTERN.findall(value):
                candidate = match.strip(" /.,")
                if "@" in candidate:
                    continue
                if not candidate.startswith(("http://", "https://")):
                    candidate = f"https://{candidate}"
                return candidate

            email = self._extract_email(value)
            if email:
                return f"https://{email.split('@', maxsplit=1)[1]}"
        return None

    def _stringify(self, value: Any) -> str | None:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, date):
            return value.isoformat()
        text = str(value).strip()
        return text or None
