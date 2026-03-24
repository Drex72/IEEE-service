from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from app.services.excel import ExcelSponsorTrackerParser


def build_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sponsor Outreach Tracker"
    sheet["A1"] = "IES UNILAG 2026"
    headers = [
        "#",
        "Company",
        "Category",
        "Tier",
        "Contact / Email",
        "Phone / Address",
        "How to Reach",
        "Date Sent",
        "Status",
        "Follow-Up Date",
        "Notes",
    ]
    for col, value in enumerate(headers, start=1):
        sheet.cell(row=4, column=col, value=value)
    sheet.cell(row=6, column=1, value=1)
    sheet.cell(row=6, column=2, value="Arnergy Solar Ltd")
    sheet.cell(row=6, column=3, value="Energy / IoT Solar")
    sheet.cell(row=6, column=4, value="Gold")
    sheet.cell(row=6, column=5, value="info@arnergy.com")
    sheet.cell(row=6, column=7, value="Email partnerships")
    sheet.cell(row=6, column=9, value="📧 Not Sent")
    sheet.cell(
        row=6,
        column=11,
        value="IoT-enabled solar company with strong energy alignment.",
    )

    summary = workbook.create_sheet("Summary & Instructions")
    summary["A1"] = "HOW TO USE THIS TRACKER"
    summary["A2"] = "STEP 1"
    summary["B2"] = "Start with high-fit companies."
    summary["A9"] = "TIER GUIDE"
    summary["B9"] = "Gold, Silver, Bronze"
    summary["A10"] = "IEEE ANGLE"
    summary["B10"] = "Lead with IEEE credibility."

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_parser_extracts_company_rows() -> None:
    parser = ExcelSponsorTrackerParser()
    companies, summary = parser.parse(build_workbook())

    assert len(companies) == 1
    assert companies[0].name == "Arnergy Solar Ltd"
    assert companies[0].contact_email == "info@arnergy.com"
    assert companies[0].website == "https://arnergy.com"
    assert len(companies[0].contacts) == 1
    assert companies[0].contacts[0].email == "info@arnergy.com"
    assert summary is not None
    assert summary.ieee_angle == "Lead with IEEE credibility."


def test_parser_classifies_generic_routes_and_contact_forms() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sponsor Outreach Tracker"
    headers = [
        "#",
        "Company",
        "Category",
        "Tier",
        "Contact / Email",
        "Phone / Address",
        "How to Reach",
        "Date Sent",
        "Status",
        "Follow-Up Date",
        "Notes",
    ]
    for col, value in enumerate(headers, start=1):
        sheet.cell(row=4, column=col, value=value)

    sheet.cell(row=6, column=2, value="ABB Nigeria")
    sheet.cell(row=6, column=5, value="new.abb.com/contact/nigeria")
    sheet.cell(row=6, column=7, value="Use ABB contact form; request Nigeria CSR team")

    sheet.cell(row=7, column=2, value="Wiretooth Technologies")
    sheet.cell(row=7, column=5, value="Via website: wiretoothtech.com or LinkedIn")
    sheet.cell(row=7, column=7, value="Reach via LinkedIn or their website contact form")

    sheet.cell(row=8, column=2, value="Example Partner")
    sheet.cell(row=8, column=5, value="Jane Doe, Partnerships Lead / jane@example.com")
    sheet.cell(row=8, column=7, value="Email Jane directly")

    buffer = BytesIO()
    workbook.save(buffer)

    parser = ExcelSponsorTrackerParser()
    companies, _summary = parser.parse(buffer.getvalue())

    abb = companies[0]
    assert abb.contact_email is None
    assert abb.contact_details == "Website contact form"
    assert abb.contacts[0].raw_contact == "Website contact form"
    assert abb.contacts[0].metadata["contact_kind"] == "contact_form"

    wiretooth = companies[1]
    assert wiretooth.contact_email is None
    assert wiretooth.contact_details == "LinkedIn or website outreach"
    assert wiretooth.contacts[0].raw_contact == "LinkedIn or website outreach"
    assert wiretooth.contacts[0].metadata["contact_kind"] == "route"

    named = companies[2]
    assert named.contact_email == "jane@example.com"
    assert named.contact_details == "Jane Doe, Partnerships Lead"
    assert named.contacts[0].full_name == "Jane Doe"
